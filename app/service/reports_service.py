import io
from datetime import datetime, timedelta
from typing import Dict, List, Set
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.repositories.reports_repo import ReportsRepository
from app.models.delivery import Delivery
from app.models.shipment import Shipment


class ReportsService:
    def __init__(self, repo: ReportsRepository):
        self.repo = repo

    async def generate_monthly_report(self, year: int, warehouse_id: str, months: List[int] = None) -> bytes:
        """Генерация Excel отчета по месяцам для конкретного склада"""
        
        # Получаем все данные за год для конкретного склада
        deliveries, shipments = await self.repo.get_operations_by_year_and_warehouse(year, warehouse_id)
        
        # Если месяцы не указаны - берем все от 1 до 12
        if months is None:
            months = list(range(1, 13))
        
        # Создаем Excel книгу
        wb = Workbook()
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        for month in months:
            await self._create_month_sheet(wb, year, month, warehouse_id, deliveries, shipments)
        
        # Сохраняем в bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()

    async def _create_month_sheet(self, wb: Workbook, year: int, month: int, warehouse_id: str,
                                deliveries: List[Delivery], shipments: List[Shipment]):
        """Создать лист для конкретного месяца и склада"""
        
        # Создаем лист с названием месяца
        month_name = self._get_month_name(month)
        ws = wb.create_sheet(title=month_name)
        
        # Получаем операции за месяц для конкретного склада
        month_deliveries = [d for d in deliveries if d.scheduled_at.month == month and d.warehouse_id == warehouse_id]
        month_shipments = [s for s in shipments if s.scheduled_at.month == month and s.warehouse_id == warehouse_id]
        
        # Получаем уникальных поставщиков и дни месяца
        suppliers = self._get_unique_suppliers(month_deliveries, month_shipments)
        days_in_month = self._get_days_in_month(year, month)
        
        # Создаем заголовки
        self._create_headers(ws, days_in_month, warehouse_id)
        
        # Заполняем данные по поставщикам
        self._fill_supplier_data(ws, suppliers, days_in_month, month_deliveries, month_shipments)
        
        # Настраиваем стили
        self._apply_styles(ws, len(suppliers), len(days_in_month))

    def _get_month_name(self, month: int) -> str:
        """Получить название месяца"""
        months = [
            'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
            'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
        ]
        return months[month - 1]

    def _get_unique_suppliers(self, deliveries: List[Delivery], shipments: List[Shipment]) -> List[str]:
        """Получить уникальных поставщиков и заказчиков"""
        suppliers = set()
        
        for delivery in deliveries:
            if delivery.supplier:
                suppliers.add(delivery.supplier)
        
        for shipment in shipments:
            if shipment.customer:
                suppliers.add(shipment.customer)
        
        return sorted(list(suppliers))

    def _get_days_in_month(self, year: int, month: int) -> List[int]:
        """Получить список дней в месяце"""
        if month == 12:
            next_month = 1
            next_year = year + 1
        else:
            next_month = month + 1
            next_year = year
        
        first_day = datetime(year, month, 1)
        last_day = datetime(next_year, next_month, 1) - timedelta(days=1)
        
        return list(range(1, last_day.day + 1))

    def _create_headers(self, ws, days_in_month, warehouse_id: str):
        """Создать заголовки таблицы"""
        # Заголовок месяца и склада
        ws.merge_cells('A1:B1')
        ws['A1'] = f'Поставщик/День - Склад {warehouse_id}'
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Заголовки дней
        for col, day in enumerate(days_in_month, 3):  # Начинаем с колонки C
            cell = ws.cell(row=1, column=col, value=day)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _fill_supplier_data(self, ws, suppliers, days_in_month, deliveries, shipments):
        """Заполнить данные по поставщикам"""
        
        # Цвета для ячеек
        delivery_fill = PatternFill(start_color='BB80FF', end_color='BB80FF', fill_type='solid')  # Фиолетовый для поставок
        shipment_fill = PatternFill(start_color='FFA789', end_color='FFA789', fill_type='solid')   # Оранжевый для отгрузок
        both_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')       # Розовый для обоих
        
        for row, supplier in enumerate(suppliers, 2):  # Начинаем со строки 2
            ws.cell(row=row, column=1, value=supplier)
            
            for col, day in enumerate(days_in_month, 3):  # Начинаем с колонки C
                cell = ws.cell(row=row, column=col)
                
                # Проверяем поставки
                supplier_deliveries = [
                    d for d in deliveries 
                    if d.supplier == supplier and d.scheduled_at.day == day
                ]
                
                # Проверяем отгрузки
                supplier_shipments = [
                    s for s in shipments 
                    if s.customer == supplier and s.scheduled_at.day == day
                ]
                
                # Заполняем ячейку
                if supplier_deliveries and supplier_shipments:
                    cell.value = 'ДТ+ОТ'
                    cell.fill = both_fill
                elif supplier_deliveries:
                    cell.value = 'ДТ'
                    cell.fill = delivery_fill
                elif supplier_shipments:
                    cell.value = 'ОТ'
                    cell.fill = shipment_fill

    def _apply_styles(self, ws, suppliers_count, days_count):
        """Применить стили к таблице"""
        
        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем границы ко всем ячейкам
        for row in range(1, suppliers_count + 2):
            for col in range(1, days_count + 3):
                ws.cell(row=row, column=col).border = thin_border
        
        # Настраиваем ширину колонок
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 5
        
        for col in range(3, days_count + 3):
            ws.column_dimensions[get_column_letter(col)].width = 8
        
        # Жирный шрифт для заголовков
        for col in range(1, days_count + 3):
            ws.cell(row=1, column=col).font = Font(bold=True)